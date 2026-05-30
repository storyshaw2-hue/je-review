"""
report.py — Write the JE review workpaper.

This documents a *review*, not just a list of flags. The automated checks are a
first-pass screen (Triage checks) that surface entries to look at; the reviewer
records, per entry, whether support agrees on amount / date / account /
description, a disposition, and a conclusion.

Workbook sheets:
  Review Summary — run metadata, population shape, triage-check results,
                   review status (disposition breakdown + support coverage), Benford
  Review Queue   — one row per surfaced entry: why it surfaced + the reviewer's
                   support review (assertions, disposition, correction, conclusion)
  Support Log    — one row per support item attached to an entry (metadata only)
  Triage checks  — every check, its weight/category, whether it ran (+ scorecard)
  Triage Notes   — plain-English notes for the top entries (optional)

Also writes a flat CSV mirroring the Review Queue. No support-file binaries are
stored — only metadata and conclusions.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .engine import RunResult
from .review import (
    ReviewRecord, DISPOSITIONS, SUPPORT_STATUSES, ASSERTION_OUTCOMES, ASSERTIONS,
    ACCEPTED_DISPOSITIONS, ISSUE_DISPOSITIONS, PENDING_DISPOSITIONS,
)

REVIEW_SHEET = "Review Queue"

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
SUBHEAD_FONT = Font(name=FONT, bold=True, size=11, color="1F3864")
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="595959")
BASE_FONT = Font(name=FONT, size=10)
MUTED_FONT = Font(name=FONT, size=10, color="9A9A9A")

# Disposition palette (keys must match review.DISPOSITIONS)
DISP_FILLS = {
    "Open":                         PatternFill("solid", fgColor="FFF4D6"),  # amber
    "Recorded correctly":           PatternFill("solid", fgColor="E2EFDA"),  # green
    "Expected business activity":   PatternFill("solid", fgColor="E2EFDA"),  # green
    "Needs support / explanation":  PatternFill("solid", fgColor="FFF0CC"),  # light amber
    "Needs correction":             PatternFill("solid", fgColor="FCE4D6"),  # red-orange
    "False positive":              PatternFill("solid", fgColor="EDEDED"),  # grey
    "Escalate":                     PatternFill("solid", fgColor="F4B5B5"),  # red
}
ASSERT_FILLS = {
    "Agrees":        PatternFill("solid", fgColor="E2EFDA"),
    "Discrepancy":   PatternFill("solid", fgColor="F4B5B5"),
    "Not in support": PatternFill("solid", fgColor="FFF0CC"),
}
SKIPPED_ROW_FILL = PatternFill("solid", fgColor="F5F5F5")


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")


def _priority_band(score) -> str:
    try:
        s = float(score)
    except (TypeError, ValueError):
        return ""
    if s >= 5:
        return "High"
    if s >= 3:
        return "Medium"
    if s >= 1:
        return "Low"
    return ""


def _rec(reviews, je) -> ReviewRecord | None:
    """Return a ReviewRecord for an entry, coercing dicts (incl. the legacy
    {disposition, note} shape) so older callers keep working."""
    if not reviews:
        return None
    v = reviews.get(str(je)) if str(je) in reviews else reviews.get(je)
    if v is None:
        return None
    if isinstance(v, ReviewRecord):
        return v
    d = dict(v)
    if "conclusion" not in d and "note" in d:
        d["conclusion"] = d.get("note", "")
    d.setdefault("je_id", str(je))
    return ReviewRecord.from_dict(d)


def _add_dropdown(ws, col_letter, max_row, choices, title):
    if max_row < 2:
        return
    dv = DataValidation(type="list", formula1='"' + ",".join(choices) + '"',
                        allow_blank=True, showDropDown=False, errorStyle="warning")
    dv.error = "Pick one of: " + ", ".join(choices)
    dv.errorTitle = title
    dv.prompt = title
    dv.promptTitle = title
    dv.add(f"{col_letter}2:{col_letter}{max_row}")
    ws.add_data_validation(dv)


def write_excel(result: RunResult, out_path, source_name: str = "",
                triage_md: str | None = None, reviews: dict | None = None,
                scorecard: dict | None = None):
    target = out_path if hasattr(out_path, "write") else Path(out_path)
    wb = Workbook()
    df_entries = result.entries.copy() if not result.entries.empty else pd.DataFrame()

    # =========================================================
    # Review Summary
    # =========================================================
    ws = wb.active
    ws.title = "Review Summary"
    ws["A1"] = "Journal Entry Review"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    meta = [
        ("Source file", source_name or "(in-memory)"),
        ("Run timestamp", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Period(s)", result.stats.get("periods", "")),
        ("Total entries", result.stats.get("total_entries", 0)),
        ("Total lines", result.stats.get("total_lines", 0)),
        ("Surfaced for review", result.stats.get("flagged_entries", 0)),
        ("Share surfaced", f"{result.stats.get('flagged_pct', 0)}%"),
        ("Likely recording errors", result.stats.get("accuracy_entries", 0)),
        ("Anomalies to confirm", result.stats.get("risk_entries", 0)),
    ]
    r = 3
    for k, v in meta:
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=r, column=2, value=v).font = BASE_FONT
        r += 1

    # --- Review status (NEW) ---
    r += 1
    ws.cell(row=r, column=1, value="Review status").font = SUBHEAD_FONT
    r += 1
    flagged_ids = list(df_entries["je_id"]) if not df_entries.empty else []
    recs = [_rec(reviews, j) for j in flagged_ids]
    recs = [x for x in recs if x is not None]
    disp_counts = {d: 0 for d in DISPOSITIONS}
    support_reviewed = 0
    for rc in recs:
        disp_counts[rc.disposition] = disp_counts.get(rc.disposition, 0) + 1
        if rc.support_reviewed:
            support_reviewed += 1
    reviewed = sum(v for d, v in disp_counts.items() if d != "Open")
    n_flagged = len(flagged_ids)
    status_rows = [
        ("Entries with a disposition", f"{reviewed} of {n_flagged}"),
        ("Support reviewed / not required", f"{support_reviewed} of {n_flagged}"),
    ]
    for d in DISPOSITIONS:
        if disp_counts.get(d):
            status_rows.append((f"  {d}", disp_counts[d]))
    for k, v in status_rows:
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=k.strip() in ("Entries with a disposition", "Support reviewed / not required"), size=10)
        ws.cell(row=r, column=2, value=v).font = BASE_FONT
        r += 1

    # --- Population shape ---
    r += 1
    ws.cell(row=r, column=1, value="Population shape").font = SUBHEAD_FONT
    r += 1
    for k, v in _population_shape(df_entries, result):
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        cell = ws.cell(row=r, column=2, value=v)
        cell.font = BASE_FONT
        if isinstance(v, (int, float)) and not isinstance(v, bool) and "amount" in k.lower():
            cell.number_format = "$#,##0;($#,##0);-"
        r += 1

    # --- Triage-check results ---
    r += 1
    ws.cell(row=r, column=1, value="First-pass checks (triage)").font = SUBHEAD_FONT
    r += 1
    hdr = ["Check", "Weight", "Ran?", "Entries surfaced", "% of surfaced", "What it needs"]
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, len(hdr))
    r += 1
    rs = result.rule_summary
    flagged_total = int(result.stats.get("flagged_entries", 0)) or 0
    for _, row in rs.iterrows():
        ran = bool(row["ran"])
        label_cell = ws.cell(row=r, column=1, value=row["label"])
        weight_cell = ws.cell(row=r, column=2, value=int(row["weight"]))
        ran_cell = ws.cell(row=r, column=3, value="Yes" if ran else "No — field missing")
        count_cell = ws.cell(row=r, column=4,
                             value=f"=COUNTIF('{REVIEW_SHEET}'!F:F,\"*\"&\"{row['rule_id']}\"&\"*\")")
        pct_cell = ws.cell(row=r, column=5,
                           value=(f"=IFERROR(D{r}/{flagged_total},0)" if flagged_total > 0 else '=""'))
        pct_cell.number_format = "0.0%"
        needs_cell = ws.cell(row=r, column=6, value=row["requires"])
        cells = (label_cell, weight_cell, ran_cell, count_cell, pct_cell, needs_cell)
        for c in cells:
            c.font = BASE_FONT if ran else MUTED_FONT
            if not ran:
                c.fill = SKIPPED_ROW_FILL
        for c in (weight_cell, ran_cell, count_cell, pct_cell):
            c.alignment = Alignment(horizontal="center")
        r += 1

    # --- Benford ---
    r += 1
    b = result.benford or {}
    if b.get("mad") is not None:
        ws.cell(row=r, column=1, value="Benford first-digit diagnostic (population)").font = SUBHEAD_FONT
        r += 1
        ws.cell(row=r, column=1, value=f"Mean abs. deviation (MAD): {b['mad']}  →  {b['verdict']}").font = BASE_FONT
        r += 1
        ws.cell(row=r, column=1,
                value=f"Based on {b['n']} line amounts ≥ 1. MAD bands per Nigrini: "
                      "<0.006 close, <0.012 acceptable, <0.015 marginal, else investigate.").font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    else:
        msg = b.get("verdict") or "Too few amounts for a reliable Benford read"
        ws.cell(row=r, column=1, value=f"Benford diagnostic: {msg} (need ≥ 50 amounts).").font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # --- Disclaimer ---
    r += 2
    disclaimer = ("This workbook documents a review. The automated checks are a first-pass screen "
                  "that surfaces entries to review — they are not conclusions. A reviewer confirms "
                  "whether each entry was recorded correctly using support. Runs entirely locally; "
                  "no ledger or support data leaves this machine.")
    ws.cell(row=r, column=1, value=disclaimer).font = NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 60
    for col, w in {"A": 32, "B": 14, "C": 20, "D": 16, "E": 14, "F": 42}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # =========================================================
    # Review Queue
    # =========================================================
    we = wb.create_sheet(REVIEW_SHEET)
    # Check IDs MUST stay on column F so the Summary COUNTIF resolves.
    order = ["je_id", "risk_score", "__band", "tests_fired", "reasons", "test_ids",
             "entry_date", "period", "lines", "total_debit", "total_credit",
             "net_amount", "entered_by", "approved_by", "source", "description"]
    nice = ["JE ID", "Review priority", "Priority", "# Checks", "Why surfaced", "Check IDs",
            "Entry Date", "Period", "Lines", "Total Debit", "Total Credit",
            "Net", "Preparer", "Approver", "Source", "Description"]
    review_cols = ["__type", "__disp", "__supstat", "__amt", "__date", "__acct",
                   "__desc", "__support", "__correction", "__reviewer", "__revdate", "__concl"]
    review_nice = ["Flag Type", "Disposition", "Support Status", "Amount?", "Date?",
                   "Account?", "Description?", "Support", "Correction", "Reviewer",
                   "Review Date", "Conclusion"]
    full_order = order + review_cols
    full_nice = nice + review_nice

    df = df_entries if not df_entries.empty else pd.DataFrame(columns=[c for c in order if not c.startswith("__")])
    df = df.assign(__band=df.get("risk_score", pd.Series([], dtype=float)).map(_priority_band)
                   if "risk_score" in df.columns else "")
    for c in order:
        if c not in df.columns:
            df[c] = ""
    df["__type"] = df["flag_type"] if "flag_type" in df.columns else ""

    # Pull review fields per entry
    def _f(je, attr, default=""):
        rc = _rec(reviews, je)
        if rc is None:
            return default
        if attr in ASSERTIONS:
            return rc.assertions.get(attr, "")
        if attr == "support":
            return "; ".join(f"{si.label} ({si.support_type})" for si in rc.support_items if si.label)
        return getattr(rc, attr, default)

    df["__disp"] = df["je_id"].map(lambda j: _f(j, "disposition", "Open"))
    df["__supstat"] = df["je_id"].map(lambda j: _f(j, "support_status", ""))
    df["__amt"] = df["je_id"].map(lambda j: _f(j, "amount"))
    df["__date"] = df["je_id"].map(lambda j: _f(j, "date"))
    df["__acct"] = df["je_id"].map(lambda j: _f(j, "account"))
    df["__desc"] = df["je_id"].map(lambda j: _f(j, "description"))
    df["__support"] = df["je_id"].map(lambda j: _f(j, "support"))
    df["__correction"] = df["je_id"].map(lambda j: _f(j, "correction"))
    df["__reviewer"] = df["je_id"].map(lambda j: _f(j, "reviewer"))
    df["__revdate"] = df["je_id"].map(lambda j: _f(j, "review_date"))
    df["__concl"] = df["je_id"].map(lambda j: _f(j, "conclusion"))

    we.append(full_nice)
    _style_header(we, 1, len(full_nice))
    for _, row in df.iterrows():
        vals = []
        for c in full_order:
            v = row.get(c, "")
            if c == "entry_date" and pd.notna(v) and v != "":
                try:
                    v = pd.Timestamp(v).strftime("%Y-%m-%d")
                except Exception:
                    pass
            vals.append("" if (isinstance(v, float) and pd.isna(v)) or v is None else v)
        we.append(vals)

    band_i = full_order.index("__band")
    type_i = full_order.index("__type")
    disp_i = full_order.index("__disp")
    triage_je_rows = _index_triage_je_rows(triage_md) if triage_md else {}
    wrap_fields = {"reasons", "description", "__concl", "__support", "__correction"}

    for rr in range(2, we.max_row + 1):
        for cc in range(1, len(full_nice) + 1):
            cell = we.cell(row=rr, column=cc)
            cell.font = BASE_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(full_order[cc - 1] in wrap_fields))
        for col_field in ("total_debit", "total_credit", "net_amount"):
            we.cell(row=rr, column=full_order.index(col_field) + 1).number_format = "#,##0.00;(#,##0.00);-"
        for ci in (band_i, type_i):
            we.cell(row=rr, column=ci + 1).alignment = Alignment(horizontal="center", vertical="top")
        if triage_md:
            je_cell = we.cell(row=rr, column=full_order.index("je_id") + 1)
            if je_cell.value and str(je_cell.value) in triage_je_rows:
                je_cell.hyperlink = f"#Triage!A{triage_je_rows[str(je_cell.value)]}"
                je_cell.font = Font(name=FONT, size=10, color="0563C1", underline="single")

    # Conditional formatting
    if we.max_row >= 2:
        last = we.max_row
        we.conditional_formatting.add(
            f"B2:B{last}", ColorScaleRule(start_type="min", start_color="FFFFFF",
                                          end_type="max", end_color="C00000"))
        band_L = get_column_letter(band_i + 1)
        for val, fg, fcol in (("High", "F4B5B5", "9C1B1B"), ("Medium", "FFE4A8", "9C6A1B"),
                              ("Low", "E2EFDA", "375623")):
            we.conditional_formatting.add(
                f"{band_L}2:{band_L}{last}",
                CellIsRule(operator="equal", formula=[f'"{val}"'],
                           fill=PatternFill("solid", fgColor=fg),
                           font=Font(name=FONT, size=10, bold=(val == "High"), color=fcol)))
        type_L = get_column_letter(type_i + 1)
        for val, fg in (("Accuracy", "DDEBF7"), ("Risk", "FFF4D6"), ("Both", "FCE4D6")):
            we.conditional_formatting.add(
                f"{type_L}2:{type_L}{last}",
                CellIsRule(operator="equal", formula=[f'"{val}"'], fill=PatternFill("solid", fgColor=fg)))
        # Disposition dropdown + fills
        disp_L = get_column_letter(disp_i + 1)
        _add_dropdown(we, disp_L, last, DISPOSITIONS, "Disposition")
        for choice, fill in DISP_FILLS.items():
            we.conditional_formatting.add(
                f"{disp_L}2:{disp_L}{last}",
                CellIsRule(operator="equal", formula=[f'"{choice}"'], fill=fill))
        # Support-status dropdown
        _add_dropdown(we, get_column_letter(full_order.index("__supstat") + 1), last,
                      SUPPORT_STATUSES, "Support status")
        # Assertion dropdowns + fills
        for key in ("__amt", "__date", "__acct", "__desc"):
            col_L = get_column_letter(full_order.index(key) + 1)
            _add_dropdown(we, col_L, last, ASSERTION_OUTCOMES, "Support agrees?")
            for val, fill in ASSERT_FILLS.items():
                we.conditional_formatting.add(
                    f"{col_L}2:{col_L}{last}",
                    CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill))

    widths = {"JE ID": 15, "Review priority": 11, "Priority": 9, "# Checks": 8,
              "Why surfaced": 52, "Check IDs": 20, "Entry Date": 12, "Period": 9,
              "Lines": 6, "Total Debit": 13, "Total Credit": 13, "Net": 12,
              "Preparer": 12, "Approver": 12, "Source": 9, "Description": 26,
              "Flag Type": 10, "Disposition": 22, "Support Status": 22,
              "Amount?": 11, "Date?": 11, "Account?": 11, "Description?": 12,
              "Support": 34, "Correction": 26, "Reviewer": 12, "Review Date": 12,
              "Conclusion": 40}
    for i, h in enumerate(full_nice, start=1):
        we.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)
    we.freeze_panes = "B2"
    we.auto_filter.ref = f"A1:{get_column_letter(len(full_nice))}{max(we.max_row, 1)}"

    # =========================================================
    # Support Log
    # =========================================================
    wl = wb.create_sheet("Support Log")
    lhead = ["JE ID", "Support", "Type", "Covers", "Agrees?", "Note"]
    wl.append(lhead)
    _style_header(wl, 1, len(lhead))
    any_support = False
    for je in flagged_ids:
        rc = _rec(reviews, je)
        if not rc:
            continue
        for si in rc.support_items:
            if not si.label and not si.note:
                continue
            any_support = True
            wl.append([je, si.label, si.support_type, ", ".join(si.covers), si.agrees, si.note])
    if not any_support:
        wl.append(["—", "No support items recorded yet.", "", "", "", ""])
        wl.cell(row=2, column=2).font = MUTED_FONT
    for rr in range(2, wl.max_row + 1):
        for cc in range(1, len(lhead) + 1):
            wl.cell(row=rr, column=cc).font = BASE_FONT
            wl.cell(row=rr, column=cc).alignment = Alignment(vertical="top", wrap_text=(cc == 6))
    if not any_support:
        wl.cell(row=2, column=2).font = MUTED_FONT
    for col, w in {"A": 15, "B": 34, "C": 18, "D": 22, "E": 14, "F": 50}.items():
        wl.column_dimensions[col].width = w
    wl.freeze_panes = "A2"

    # =========================================================
    # Triage checks
    # =========================================================
    wc = wb.create_sheet("Triage checks")
    use_score = bool(scorecard)
    chead = ["Check ID", "Check", "Category", "Weight", "Fields needed", "Ran?", "Description"]
    if use_score:
        chead += ["Surfaced", "Benign", "Issues", "Signal"]
    wc.append(chead)
    _style_header(wc, 1, len(chead))
    for _, row in result.rule_summary.iterrows():
        ran = bool(row["ran"])
        rowvals = [row["rule_id"], row["label"], row.get("category", "Risk"), int(row["weight"]),
                   row["requires"], "Yes" if ran else "No", row["description"]]
        if use_score:
            sc = scorecard.get(row["rule_id"], {})
            rowvals += [sc.get("surfaced", ""), sc.get("accepted", ""), sc.get("issue", ""),
                        sc.get("recommendation", "")]
        wc.append(rowvals)
        rr = wc.max_row
        for cc in range(1, len(chead) + 1):
            cell = wc.cell(row=rr, column=cc)
            cell.font = BASE_FONT if ran else MUTED_FONT
            if not ran:
                cell.fill = SKIPPED_ROW_FILL
            cell.alignment = Alignment(vertical="top", wrap_text=(cc in (7, len(chead))))
    cwidths = {"A": 24, "B": 26, "C": 11, "D": 8, "E": 22, "F": 7, "G": 60}
    if use_score:
        cwidths.update({"H": 10, "I": 9, "J": 8, "K": 34})
    for col, w in cwidths.items():
        wc.column_dimensions[col].width = w
    wc.freeze_panes = "A2"

    # =========================================================
    # Triage Notes (optional)
    # =========================================================
    if triage_md:
        wt = wb.create_sheet("Triage")
        wt["A1"] = "Triage Notes"
        wt["A1"].font = TITLE_FONT
        rownum = 3
        for line in triage_md.splitlines():
            cell = wt.cell(row=rownum, column=1, value=line)
            if line.startswith("## "):
                cell.value = line[3:]
                cell.font = Font(name=FONT, bold=True, size=11, color="1F3864")
            elif line.startswith("# "):
                cell.value = line[2:]
                cell.font = Font(name=FONT, bold=True, size=12)
            else:
                cell.font = BASE_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            rownum += 1
        wt.column_dimensions["A"].width = 110
        wt.freeze_panes = "A2"

    wb.save(target)
    return target


def _population_shape(df_entries: pd.DataFrame, result: RunResult) -> list:
    rows = []
    stats = result.stats or {}
    if not df_entries.empty:
        if "entered_by" in df_entries.columns:
            rows.append(("Unique preparers (surfaced)", int(df_entries["entered_by"].nunique())))
        if "approved_by" in df_entries.columns:
            rows.append(("Unique approvers (surfaced)", int(df_entries["approved_by"].nunique())))
        if "entry_date" in df_entries.columns:
            try:
                dates = pd.to_datetime(df_entries["entry_date"], errors="coerce").dropna()
                if not dates.empty:
                    rows.append(("Earliest surfaced date", dates.min().strftime("%Y-%m-%d")))
                    rows.append(("Latest surfaced date", dates.max().strftime("%Y-%m-%d")))
            except Exception:
                pass
        for amt_col in ("total_debit", "net_amount"):
            if amt_col in df_entries.columns:
                try:
                    series = pd.to_numeric(df_entries[amt_col], errors="coerce").dropna().abs()
                    if not series.empty:
                        label = "Total surfaced $ (debits)" if amt_col == "total_debit" else "Total surfaced net $"
                        rows.append((label, round(float(series.sum()), 2)))
                        break
                except Exception:
                    pass
    for key in ("avg_entry_amount", "max_entry_amount"):
        if key in stats:
            rows.append((key.replace("_", " ").capitalize(), stats[key]))
    if not rows:
        rows.append(("(no surfaced entries)", "—"))
    return rows


def _index_triage_je_rows(triage_md: str) -> dict:
    je_rows: dict[str, int] = {}
    rownum = 3
    for line in triage_md.splitlines():
        stripped = line.strip()
        if "  ·  risk " in stripped or " · risk " in stripped:
            cleaned = stripped.lstrip("#").strip()
            je_id = cleaned.split()[0] if cleaned else ""
            if je_id and je_id not in je_rows:
                je_rows[je_id] = rownum
        rownum += 1
    return je_rows


def write_csv(result: RunResult, out_path, reviews: dict | None = None):
    target = out_path if hasattr(out_path, "write") else Path(out_path)
    df = result.entries.copy()
    if "risk_score" in df.columns:
        df.insert(df.columns.get_loc("risk_score") + 1, "priority_band",
                  df["risk_score"].map(_priority_band))
    if reviews:
        def g(j, attr):
            rc = _rec(reviews, j)
            if rc is None:
                return ""
            if attr in ASSERTIONS:
                return rc.assertions.get(attr, "")
            return getattr(rc, attr, "")
        df["disposition"] = df["je_id"].map(lambda j: g(j, "disposition"))
        df["support_status"] = df["je_id"].map(lambda j: g(j, "support_status"))
        for a in ASSERTIONS:
            df[f"{a}_agrees"] = df["je_id"].map(lambda j, a=a: g(j, a))
        df["correction"] = df["je_id"].map(lambda j: g(j, "correction"))
        df["reviewer"] = df["je_id"].map(lambda j: g(j, "reviewer"))
        df["review_date"] = df["je_id"].map(lambda j: g(j, "review_date"))
        df["conclusion"] = df["je_id"].map(lambda j: g(j, "conclusion"))
    df.to_csv(target, index=False)
    return target
